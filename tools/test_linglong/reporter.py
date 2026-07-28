#!/usr/bin/env python3
# SPDX-FileCopyrightText: 2026 UnionTech Software Technology Co., Ltd.
#
# SPDX-License-Identifier: LGPL-3.0-or-later

"""
玲珑冒烟测试 — 报告生成

支持 JSON 和 Excel 两种输出格式。
"""

import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from .models import StepResult

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_report(
    results: list[StepResult],
    start_time: datetime,
    results_file: str,
    dated: bool = False,
) -> None:
    """生成测试报告（JSON + Excel）。"""
    total = len(results)
    passed = sum(1 for r in results if r.status == "PASS")
    failed = sum(1 for r in results if r.status == "FAIL")
    skipped = sum(1 for r in results if r.status == "SKIPPED")
    duration_sec = sum(r.duration_ms for r in results) / 1000.0

    # ── JSON ──
    summary = {
        "start_time": start_time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "total_steps": total,
        "passed": passed,
        "failed": failed,
        "skipped": skipped,
        "duration_seconds": duration_sec,
        "steps": [
            {
                "index": r.index,
                "title": r.title,
                "status": r.status,
                "duration_ms": r.duration_ms,
                "error_message": r.error_message,
            }
            for r in results
        ],
    }
    with open(results_file, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"\nJSON results saved: {results_file}")

    # ── Excel ──
    if not HAS_OPENPYXL:
        print(
            "Warning: openpyxl not available, skipping Excel report",
            file=sys.stderr,
        )
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "玲珑冒烟测试"

    # 样式
    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pass_font = Font(name="Microsoft YaHei", size=10, color="375623", bold=True)
    fail_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    fail_font = Font(name="Microsoft YaHei", size=10, color="C62828", bold=True)
    skip_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    skip_font = Font(name="Microsoft YaHei", size=10, color="E65100", bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, w in {1: 8, 2: 14, 3: 50, 4: 14, 5: 12, 6: 40}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # 标题行
    ws.merge_cells("A1:F1")
    ws["A1"].value = "玲珑冒烟测试报告"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # 汇总行
    ws.merge_cells("A2:F2")
    start_time_str = start_time.strftime("%Y-%m-%d %H:%M:%S")
    ws["A2"].value = (
        f"测试时间: {start_time_str}  |  总计: {total}  |  "
        f"通过: {passed}  |  失败: {failed}  |  跳过: {skipped}  |  "
        f"运行时长: {duration_sec:.1f}s"
    )
    ws["A2"].font = Font(name="Microsoft YaHei", size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 26

    # 表头
    for ci, h in enumerate(
        ["序号", "模块", "测试用例", "执行状态", "耗时(ms)", "错误信息"], 1
    ):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
    ws.row_dimensions[3].height = 28

    # 数据行
    for i, r in enumerate(results):
        row = i + 4
        ws.row_dimensions[row].height = 30
        for ci, val in [
            (1, r.index),
            (2, r.category or ""),
            (3, r.title),
            (4, r.status),
            (5, r.duration_ms),
            (6, r.error_message or ""),
        ]:
            c = ws.cell(row=row, column=ci, value=val)
            c.border = thin_border
            c.alignment = Alignment(
                horizontal="center" if ci in (1, 2, 4, 5) else "left",
                vertical="center",
            )
            c.font = Font(name="Microsoft YaHei", size=10)
            if ci == 4:
                if val == "PASS":
                    c.fill = pass_fill
                    c.font = pass_font
                elif val == "FAIL":
                    c.fill = fail_fill
                    c.font = fail_font
                elif val == "SKIPPED":
                    c.fill = skip_fill
                    c.font = skip_font

    # 输出文件名
    output_name = "玲珑冒烟测试.xlsx"
    if dated:
        p = Path(output_name)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = str(p.parent / f"{p.stem}_{ts}{p.suffix}")

    wb.save(output_name)
    print(f"Excel report saved: {output_name}")
